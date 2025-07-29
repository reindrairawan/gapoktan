from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseForbidden, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models.functions import ExtractMonth
from django.db.models import Count, Sum
from django.contrib.auth.models import User
from django.contrib import messages
from .models import ProfilGapoktan, Kegiatan, Grup, KetuaPoktan, KetuaGapoktan, Petani, Alsintan, Lahan, DataERDKK, Sppt, PenyakitTanaman
from .forms import ProfilGapoktanForm, KegiatanForm, GrupForm, KetuaPoktanForm, KetuaGapoktanForm, PetaniForm, AlsintanForm, LahanForm, DataERDKKForm, SpptForm, PenyakitTanamanForm 
from datetime import datetime
from django.utils.text import slugify
from django.core.paginator import Paginator
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.templatetags.static import static
import uuid 
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import os


# Initialize Firebase app and Firestore client if not already initialized
if not firebase_admin._apps:
    cred_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'aplikasi', 'serviceAccountKey.json')
    cred = credentials.Certificate(cred_path)
    firebase_admin.initialize_app(cred)
db = firestore.client()

def login_view(request):
    if request.method == 'POST':
        nik = request.POST.get('nik')
        print("NIK diterima dari form:", nik)

        docs = db.collection('user').where('nik', '==', nik).stream()
        user_data = None
        for doc in docs:
            user_data = doc.to_dict()
            print("User ditemukan di Firestore:", user_data)
            break

        if user_data:
            user, created = User.objects.get_or_create(username=nik)
            user.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, user)

            # Simpan data ke session
            request.session['username'] = user_data.get('username', '')
            request.session['role'] = user_data.get('role', '')

            print("Login berhasil. Redirect berdasarkan role...")

            # Redirect berdasarkan role
            if user_data.get('role') == 'admin':
                return redirect('beranda_admin')  # pastikan url name ini sesuai di urls.py
            else:
                return redirect('beranda')  # untuk user biasa
        
        messages.error(request, 'NIK tidak ditemukan.')

    return render(request, 'login.html')

@login_required(login_url='login')
def dashboard(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
   # Data petani per bulan
    petani_per_bulan = [Petani.objects.filter(created_at__month=bulan).count() for bulan in range(1, 13)]

    # Data lahan per bulan
    lahan_per_bulan = [Lahan.objects.filter(created_at__month=bulan).count() for bulan in range(1, 13)]

    # Data alsintan per bulan
    # Data alsintan per bulan
    total_alat = Alsintan.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    alat_tersedia = Alsintan.objects.filter(kondisi='tersedia').aggregate(total=Sum('jumlah'))['total'] or 0
    alat_dipinjam = Alsintan.objects.filter(kondisi='dipinjam').aggregate(total=Sum('jumlah'))['total'] or 0
    alat_rusak = Alsintan.objects.filter(kondisi='rusak').aggregate(total=Sum('jumlah'))['total'] or 0

    
    # Total kebutuhan pupuk
    total_urea = DataERDKK.objects.aggregate(total=Sum('pupuk_urea'))['total'] or 0
    total_npk = DataERDKK.objects.aggregate(total=Sum('pupuk_npk'))['total'] or 0
    total_organik = DataERDKK.objects.aggregate(total=Sum('pupuk_organik'))['total'] or 0
    
    context = {
    "judul": "Halaman Admin",
    "menu": "beranda",
    'petani_per_bulan': petani_per_bulan,
    'lahan_per_bulan': lahan_per_bulan,
    'total_alat': total_alat,
    'alat_tersedia': alat_tersedia,
    'alat_dipinjam': alat_dipinjam,
    'alat_rusak': alat_rusak,
    'urea': total_urea,
    'npk': total_npk,
    'organik': total_organik,
    }
    return render(request, 'dashboardadmin.html', context)

@login_required(login_url='login')
def unduh_laporan_pdf(request):
    # Daftar nama bulan
    bulan_list = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun',
                  'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']

    # Ambil data real dari database
    petani_per_bulan = [Petani.objects.filter(created_at__month=bulan).count() for bulan in range(1, 13)]
    
    lahan_per_bulan = [Lahan.objects.filter(created_at__month=bulan).count() for bulan in range(1, 13)]
    
    total_urea = DataERDKK.objects.aggregate(total=Sum('pupuk_urea'))['total'] or 0
    total_npk = DataERDKK.objects.aggregate(total=Sum('pupuk_npk'))['total'] or 0
    total_organik = DataERDKK.objects.aggregate(total=Sum('pupuk_organik'))['total'] or 0
    
    alat_tersedia = Alsintan.objects.filter(kondisi='tersedia').count()
    alat_dipinjam = Alsintan.objects.filter(kondisi='dipinjam').count()
    alat_rusak = Alsintan.objects.filter(kondisi='rusak').count()

    context = {
        'bulan_list': bulan_list,
        'petani_per_bulan': petani_per_bulan,
        'lahan_per_bulan': lahan_per_bulan,
        'urea': total_urea,
        'npk': total_npk,
        'organik': total_organik,
        'alat_tersedia': alat_tersedia,
        'alat_dipinjam': alat_dipinjam,
        'alat_rusak': alat_rusak,
        'tanggal_unduh': datetime.now().strftime('%d %B %Y'),
         'logo_url': request.build_absolute_uri(static('gambar/logobpp.png'))
    }

    template = get_template('laporan_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laporan_gapoktan.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Terjadi kesalahan saat membuat PDF', status=500)
    return response


@login_required(login_url='login')
def profiladmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    profil =  ProfilGapoktan.objects.all() 
    kegiatan_list = Kegiatan.objects.all() 
    context = {
            "judul": "Data Profil",
            "menu":"profil",
            "profil_list":profil,
            "kegiatan_list": kegiatan_list,
        }
    return render(request, 'profiladmin.html', context)

@login_required(login_url='login')
def formprofiladmin(request):
    if request.method == "POST":
        form = ProfilGapoktanForm(request.POST, request.FILES)  # Ambil data dari request
        if form.is_valid():  # Validasi form
            form.save()  # Simpan data ke database
            return redirect('profiladmin')  # Redirect ke halaman profil setelah menyimpan
    else:
        form = ProfilGapoktanForm()  # Tampilkan form kosong jika GET request
    context = {
        "judul": "Form ProfilGapoktan",
        "menu": "profil",
        "form": form
    }
    return render(request, 'formprofiladmin.html', context)

@login_required(login_url='login')
def editprofiladmin(request, pk):
    profil = get_object_or_404(ProfilGapoktan, id=pk)
    if request.method == "POST":
        form = ProfilGapoktanForm(request.POST, request.FILES, instance=profil)
        if form.is_valid():
            form.save()
            return redirect('profiladmin')  # Redirect ke halaman daftar profil
    else:
        form = ProfilGapoktanForm(instance=profil)
    context = {
         "judul": "Form Edit ProfilGapoktan",
        "menu": "profil",
        "form": form
    }
    return render(request, 'formprofiladmin.html', context)

@login_required(login_url='login')
def deleteprofiladmin(request, pk):
    profil = get_object_or_404(ProfilGapoktan, pk=pk)
    if request.method == 'POST':
        profil.delete()
        return redirect('profiladmin')  
    return redirect('profiladmin')

#form kegiatan

@login_required(login_url='login')
def formkegiatanadmin(request):
    if request.method == "POST":
        form = KegiatanForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('profiladmin')
    else:
        form = KegiatanForm()
    return render(request, 'formkegiatanadmin.html', {
        "judul": "Form Kegiatan",
        "menu": "profil",
        "form": form
    })

@login_required(login_url='login')
def editkegiatanadmin(request, pk):
    kegiatan = get_object_or_404(Kegiatan, id=pk)
    if request.method == "POST":
        form = KegiatanForm(request.POST, request.FILES, instance=kegiatan)
        if form.is_valid():
            form.save()
            return redirect('profiladmin')
    else:
        form = KegiatanForm(instance=kegiatan)
    return render(request, 'formkegiatanadmin.html', {
        "judul": "Edit Kegiatan",
        "menu": "profil",
        "form": form
    })

@login_required(login_url='login')
def deletekegiatanadmin(request, pk):
    kegiatan = get_object_or_404(Kegiatan, pk=pk)
    if request.method == 'POST':
        kegiatan.delete()
        return redirect('profiladmin')
    return redirect('profiladmin')

#grup
@login_required(login_url='login')
def grupadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    grup =  Grup.objects.all() 
    kegiatan_list = Kegiatan.objects.all() 
    context = {
            "judul": "Data Grup",
            "menu":"grup",
            "grup_list":grup,
            "kegiatan_list": kegiatan_list,
        }
    return render(request, 'grupadmin.html', context)

@login_required(login_url='login')
def formgrupadmin(request):
    if request.method == "POST":
        form = GrupForm(request.POST, request.FILES)  # Ambil data dari request
        if form.is_valid():  # Validasi form
            form.save()  # Simpan data ke database
            return redirect('grupadmin')  # Redirect ke halaman grup setelah menyimpan
    else:
        form = GrupForm()  # Tampilkan form kosong jika GET request
    context = {
        "judul": "Form Grup",
        "menu": "grup",
        "form": form
    }
    return render(request, 'formgrupadmin.html', context)

@login_required(login_url='login')
def editgrupadmin(request, pk):
    grup = get_object_or_404(Grup, id=pk)
    if request.method == "POST":
        form = GrupForm(request.POST, request.FILES, instance=grup)
        if form.is_valid():
            form.save()
            return redirect('grupadmin')  # Redirect ke halaman daftar grup
    else:
        form = GrupForm(instance=grup)
    context = {
         "judul": "Form Edit Grup",
        "menu": "grup",
        "form": form
    }
    return render(request, 'formgrupadmin.html', context)

@login_required(login_url='login')
def deletegrupadmin(request, pk):
    grup = get_object_or_404(Grup, pk=pk)
    if request.method == 'POST':
        grup.delete()
        return redirect('grupadmin')  
    return redirect('grupadmin')

#ketua poktan 
@login_required(login_url='login')
def ketuaadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    ketua =  KetuaPoktan.objects.all() 
    ketua_gapoktan_list = KetuaGapoktan.objects.all()
    context = {
            "judul": "Data KetuaPoktan",
            "menu":"ketua",
            "ketua_list":ketua,
            "ketua_gapoktan_list": ketua_gapoktan_list,
        }
    return render(request, 'ketuaadmin.html', context)

@login_required(login_url='login')
def formketuaadmin(request):
    if request.method == "POST":
        form = KetuaPoktanForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('ketuaadmin')
    else:
        form = KetuaPoktanForm()
    return render(request, 'formketuaadmin.html', {
        "judul": "Form KetuaPoktan",
        "menu": "ketua",
        "form": form
    })

@login_required(login_url='login')
def editketuaadmin(request, pk):
    ketua = get_object_or_404(KetuaPoktan, id=pk)
    if request.method == "POST":
        form = KetuaPoktanForm(request.POST, request.FILES, instance=ketua)
        if form.is_valid():
            form.save()
            return redirect('ketuaadmin')
        else:
            print(form.errors)  # Ini akan tampil di terminal
    else:
        form = KetuaPoktanForm(instance=ketua)
    return render(request, 'formketuaadmin.html', {
        "judul": "Edit KetuaPoktan",
        "menu": "ketua",
        "form": form
    })

@login_required(login_url='login')
def deleteketuaadmin(request, pk):
    ketua = get_object_or_404(KetuaPoktan, pk=pk)
    if request.method == 'POST':
        ketua.delete()
        return redirect('ketuaadmin')
    return redirect('ketuaadmin')

# ketua gapoktan
@login_required(login_url='login')
def formketuagapoktanadmin(request):
    if request.method == "POST":
        form = KetuaGapoktanForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('ketuaadmin')
    else:
        form = KetuaGapoktanForm()
    return render(request, 'formketuagapoktanadmin.html', {
        "judul": "Form KetuaGapoktan",
        "menu": "ketua",
        "form": form
    })
    
@login_required(login_url='login')
def editketuagapoktanadmin(request, pk):
    ketuagapoktan = get_object_or_404(KetuaGapoktan, id=pk)
    if request.method == "POST":
        form = KetuaGapoktanForm(request.POST, request.FILES, instance=ketuagapoktan)
        if form.is_valid():
            form.save()
            return redirect('ketuaadmin')
        else:
            print(form.errors)  # Ini akan tampil di terminal
    else:
        form = KetuaGapoktanForm(instance=ketuagapoktan)
    return render(request, 'formketuagapoktanadmin.html', {
        "judul": "Edit KetuaGapoktan",
        "menu": "ketua",
        "form": form
    })

@login_required(login_url='login')
def deleteketuagapoktanadmin(request, pk):
    ketuagapoktan = get_object_or_404(KetuaGapoktan, pk=pk)
    if request.method == 'POST':
        ketuagapoktan.delete()
        return redirect('ketuaadmin')
    return redirect('ketuaadmin')

# anggota
@login_required(login_url='login')
def anggotaadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    anggota =  Petani.objects.all() 
    anggota_gapoktan_list = KetuaGapoktan.objects.all()
    context = {
            "judul": "Data Petani",
            "menu":"anggota",
            "anggota_list":anggota,
            "anggota_gapoktan_list": anggota_gapoktan_list,
        }
    return render(request, 'anggotaadmin.html', context)

@login_required(login_url='login')
def formanggotaadmin(request):
    if request.method == "POST":
        form = PetaniForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('anggotaadmin')
    else:
        form = PetaniForm()
    return render(request, 'formanggotaadmin.html', {
        "judul": "Form Petani",
        "menu": "anggota",
        "form": form
    })

@login_required(login_url='login')
def editanggotaadmin(request, pk):
    anggota = get_object_or_404(Petani, id=pk)
    if request.method == "POST":
        form = PetaniForm(request.POST, request.FILES, instance=anggota)
        if form.is_valid():
            form.save()
            return redirect('anggotaadmin')
        else:
            print(form.errors)  # Ini akan tampil di terminal
    else:
        form = PetaniForm(instance=anggota)
    return render(request, 'formanggotaadmin.html', {
        "judul": "Edit Petani",
        "menu": "anggota",
        "form": form
    })

@login_required(login_url='login')
def deleteanggotaadmin(request, pk):
    anggota = get_object_or_404(Petani, pk=pk)
    if request.method == 'POST':
        anggota.delete()
        return redirect('anggotaadmin')
    return redirect('anggotaadmin')

#alsintan
@login_required(login_url='login')
def alsintanadmin(request):
    if request.session.get('role') != 'admin':
            return HttpResponseForbidden("Anda tidak memiliki akses.")
    alsintan =  Alsintan.objects.all() 
    context = {
            "judul": "Data Alsintan",
            "menu":"alsintan",
            "alsintan_list":alsintan,
        }
    return render(request, 'alsintanadmin.html', context)

@login_required(login_url='login')
def formalsintanadmin(request):
    if request.method == "POST":
        form = AlsintanForm(request.POST, request.FILES)  # Ambil data dari request
        if form.is_valid():  # Validasi form
            form.save()  # Simpan data ke database
            return redirect('alsintanadmin')  # Redirect ke halaman alsintan setelah menyimpan
    else:
        form = AlsintanForm()  # Tampilkan form kosong jika GET request
    context = {
        "judul": "Form Alsintan",
        "menu": "alsintan",
        "form": form
    }
    return render(request, 'formalsintanadmin.html', context)

@login_required(login_url='login')
def editalsintanadmin(request, pk):
    alsintan = get_object_or_404(Alsintan, id=pk)
    if request.method == "POST":
        form = AlsintanForm(request.POST, request.FILES, instance=alsintan)
        if form.is_valid():
            form.save()
            return redirect('alsintanadmin')  # Redirect ke halaman daftar alsintan
    else:
        form = AlsintanForm(instance=alsintan)
    context = {
         "judul": "Form Edit Alsintan",
        "menu": "alsintan",
        "form": form
    }
    return render(request, 'formalsintanadmin.html', context)

@login_required(login_url='login')
def deletealsintanadmin(request, pk):
    alsintan = get_object_or_404(Alsintan, pk=pk)
    if request.method == 'POST':
        alsintan.delete()
        return redirect('alsintanadmin')  
    return redirect('alsintanadmin')


#Lahan
@login_required(login_url='login')
def lahanadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    lahan =  Lahan.objects.all() 
    context = {
            "judul": "Data Lahan",
            "menu":"lahan",
            "lahan_list":lahan,
        }
    return render(request, 'lahanadmin.html', context)

@login_required(login_url='login')
def formlahanadmin(request):
    if request.method == "POST":
        form = LahanForm(request.POST, request.FILES)  # Ambil data dari request
        if form.is_valid():  # Validasi form
            form.save()  # Simpan data ke database
            return redirect('lahanadmin')  # Redirect ke halaman lahan setelah menyimpan
    else:
        form = LahanForm()  # Tampilkan form kosong jika GET request
    context = {
        "judul": "Form Lahan",
        "menu": "lahan",
        "form": form
    }
    return render(request, 'formlahanadmin.html', context)

@login_required(login_url='login')
def editlahanadmin(request, pk):
    lahan = get_object_or_404(Lahan, id=pk)
    if request.method == "POST":
        form = LahanForm(request.POST, request.FILES, instance=lahan)
        if form.is_valid():
            form.save()
            return redirect('lahanadmin')  # Redirect ke halaman daftar lahan
    else:
        form = LahanForm(instance=lahan)
    context = {
         "judul": "Form Edit Lahan",
        "menu": "lahan",
        "form": form
    }
    return render(request, 'formlahanadmin.html', context)

@login_required(login_url='login')
def deletelahanadmin(request, pk):
    lahan = get_object_or_404(Lahan, pk=pk)
    if request.method == 'POST':
        lahan.delete()
        return redirect('lahanadmin')  
    return redirect('lahanadmin')

# erdkk
@login_required(login_url='login')
def erdkkadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    erdkk_queryset = DataERDKK.objects.all().order_by('-id')
    paginator = Paginator(erdkk_queryset, 10)  # 10 data per halaman

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    sppt_list = Sppt.objects.all()
    context = {
            "judul": "Data DataERDKK",
            "menu":"erdkk",
            'erdkk_list': page_obj,
            "sppt_list": sppt_list,
        }
    return render(request, 'erdkkadmin.html', context)

@login_required(login_url='login')
def formerdkkadmin(request):
    if request.method == "POST":
        form = DataERDKKForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('erdkkadmin')
    else:
        form = DataERDKKForm()
    return render(request, 'formerdkkadmin.html', {
        "judul": "Form DataERDKK",
        "menu": "erdkk",
        "form": form
    })

@login_required(login_url='login')
def editerdkkadmin(request, pk):
    erdkk = get_object_or_404(DataERDKK, id=pk)
    if request.method == "POST":
        form = DataERDKKForm(request.POST, request.FILES, instance=erdkk)
        if form.is_valid():
            form.save()
            return redirect('erdkkadmin')
        else:
            print(form.errors)  # Ini akan tampil di terminal
    else:
        form = DataERDKKForm(instance=erdkk)
    return render(request, 'formerdkkadmin.html', {
        "judul": "Edit DataERDKK",
        "menu": "erdkk",
        "form": form
    })

@login_required(login_url='login')
def deleteerdkkadmin(request, pk):
    erdkk = get_object_or_404(DataERDKK, pk=pk)
    if request.method == 'POST':
        erdkk.delete()
        return redirect('erdkkadmin')
    return redirect('erdkkadmin')

#form kegiatan
@login_required(login_url='login')
def formspptadmin(request):
    if request.method == "POST":
        form = SpptForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('erdkkadmin')
    else:
        form = SpptForm()
    return render(request, 'formspptadmin.html', {
        "judul": "Form Sppt",
        "menu": "erdkk",
        "form": form
    })

@login_required(login_url='login')
def editspptadmin(request, pk):
    sppt = get_object_or_404(Sppt, id=pk)
    if request.method == "POST":
        form = SpptForm(request.POST, request.FILES, instance=sppt)
        if form.is_valid():
            form.save()
            return redirect('erdkkadmin')
    else:
        form = SpptForm(instance=sppt)
    return render(request, 'formspptadmin.html', {
        "judul": "Edit Sppt",
        "menu": "erdkk",
        "form": form
    })

@login_required(login_url='login')
def deletespptadmin(request, pk):
    sppt = get_object_or_404(Sppt, pk=pk)
    if request.method == 'POST':
        sppt.delete()
        return redirect('erdkkadmin')
    return redirect('erdkkadmin')

# penyakit tanaman
@login_required(login_url='login')
def penyakittanamanadmin(request):
    if request.session.get('role') != 'admin':
        return HttpResponseForbidden("Anda tidak memiliki akses.")
    penyakittanaman =  PenyakitTanaman.objects.all() 
    context = {
            "judul": "Data PenyakitTanaman",
            "menu":"penyakittanaman",
            "penyakit_list": penyakittanaman,
        }
    return render(request, 'penyakittanamanadmin.html', context)

@login_required(login_url='login')
def formpenyakittanamanadmin(request):
    if request.method == "POST":
        form = PenyakitTanamanForm(request.POST, request.FILES)  # Ambil data dari request
        if form.is_valid():  # Validasi form
            form.save()  # Simpan data ke database
            return redirect('penyakittanamanadmin')  # Redirect ke halaman penyakittanaman setelah menyimpan
    else:
        form = PenyakitTanamanForm()  # Tampilkan form kosong jika GET request
    context = {
        "judul": "Form PenyakitTanamanGapoktan",
        "menu": "penyakittanaman",
        "form": form
    }
    return render(request, 'formpenyakittanamanadmin.html', context)

@login_required(login_url='login')
def editpenyakittanamanadmin(request, pk):
    penyakittanaman = get_object_or_404(PenyakitTanaman, id=pk)
    if request.method == "POST":
        form = PenyakitTanamanForm(request.POST, request.FILES, instance=penyakittanaman)
        if form.is_valid():
            form.save()
            return redirect('penyakittanamanadmin')  # Redirect ke halaman daftar penyakittanaman
    else:
        form = PenyakitTanamanForm(instance=penyakittanaman)
    context = {
         "judul": "Form Edit PenyakitTanamanGapoktan",
        "menu": "penyakittanaman",
        "form": form
    }
    return render(request, 'formpenyakittanamanadmin.html', context)

@login_required(login_url='login')
def deletepenyakittanamanadmin(request, pk):
    penyakittanaman = get_object_or_404(PenyakitTanaman, pk=pk)
    if request.method == 'POST':
        penyakittanaman.delete()
        return redirect('penyakittanamanadmin')  
    return redirect('penyakittanamanadmin')

# Fungsi bantu: slug unik untuk Grup
def generate_unique_slug_grup(nama_grup):
    base_slug = slugify(nama_grup)
    slug = base_slug
    counter = 1

    while Grup.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1

    return slug

# Fungsi bantu: slug unik untuk Petani
def generate_unique_slug_petani(nama):
    base_slug = slugify(nama)
    slug = base_slug
    counter = 1

    while Petani.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1

    return slug

# Fungsi utama: import Excel
def import_anggota_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            total_data = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nama, nik, no_hp, nama_grup, alamat = row[:5]

                if not nama or not nik:
                    continue  # skip baris kosong

                # Cari atau buat grup dengan slug unik
                nama_grup_clean = nama_grup.strip()
                grup = Grup.objects.filter(nama=nama_grup_clean).first()
                if not grup:
                    slug_grup = generate_unique_slug_grup(nama_grup_clean)
                    grup = Grup.objects.create(nama=nama_grup_clean, slug=slug_grup)

                # Buat slug unik untuk Petani
                slug_petani = generate_unique_slug_petani(nama)

                # Simpan ke DB
                Petani.objects.create(
                    nama=nama,
                    nik=nik,
                    no_hp=no_hp if no_hp else "-",
                    grup=grup,
                    alamat=alamat if alamat else "-",
                    slug=slug_petani
                )
                total_data += 1

            messages.success(request, f'Sukses import {total_data} data!')
        except Exception as e:
            messages.error(request, f'Gagal import: {str(e)}')

    return redirect('anggotaadmin')  # pastikan sesuai dengan nama URL list kamu



def import_erdkk_excel(request):
    print("✅ VIEW import_erdkk_excel DIJALANKAN")

    if request.method == 'POST':
        print("✅ METHOD POST diterima")

        excel_file = request.FILES.get('excel_file')
        print(f"✅ File diterima: {excel_file.name if excel_file else 'TIDAK ADA'}")

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            total_data = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                print("📄 BACA ROW:", row)

                nama_petani, komoditas, luas_lahan, urea, npk, organik, tahun_rencana = row[:7]

                if not nama_petani or not komoditas:
                    print("⚠️ Baris dilewati karena kosong")
                    continue

                # Coba cari petani
                petani = Petani.objects.filter(nama__iexact=nama_petani.strip()).first()
                if not petani:
                    print(f"❌ Petani '{nama_petani}' TIDAK DITEMUKAN di database")
                    continue

                print(f"✅ Petani ditemukan: {petani.nama}, simpan ke DataERDKK...")

                DataERDKK.objects.create(
                    petani=petani,
                    komoditas=komoditas,
                    luas_lahan=luas_lahan,
                    pupuk_urea=urea,
                    pupuk_npk=npk,
                    pupuk_organik=organik or 0,
                    tahun_rencana=tahun_rencana or 2025,
                )
                total_data += 1

            messages.success(request, f'Sukses import {total_data} data ERDKK!')
            print(f"✅ {total_data} baris berhasil disimpan")
        except Exception as e:
            print("❌ ERROR:", str(e))
            messages.error(request, f'Gagal import: {str(e)}')

    return redirect('erdkkadmin')




def logoutPage(request):
    logout(request)
    return redirect('login')